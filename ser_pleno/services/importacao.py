"""
Serviço de Importação de Dados para o Desktop CustomTkinter
Suporta importação de CSV e Excel para estudantes, agendamentos e orientações
"""
import logging
import csv
import io
import os
from typing import Optional, Dict, Any, List, Callable, Tuple
from datetime import datetime
from dataclasses import dataclass, field
from enum import Enum

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from config.db_config import get_db_connection

logger = logging.getLogger(__name__)


class ImportStatus(Enum):
    """Status de uma linha de importação"""
    SUCCESS = "success"
    WARNING = "warning"
    ERROR = "error"
    SKIPPED = "skipped"


@dataclass
class ImportResult:
    """Resultado de uma importação"""
    status: ImportStatus
    row_number: int
    message: str
    data: Optional[Dict[str, Any]] = None
    original_data: Optional[Dict[str, Any]] = None


@dataclass
class ImportReport:
    """Relatório completo de uma importação"""
    total_rows: int = 0
    success_count: int = 0
    warning_count: int = 0
    error_count: int = 0
    skipped_count: int = 0
    results: List[ImportResult] = field(default_factory=list)
    started_at: datetime = field(default_factory=datetime.now)
    finished_at: Optional[datetime] = None
    
    def add_result(self, result: ImportResult):
        self.results.append(result)
        if result.status == ImportStatus.SUCCESS:
            self.success_count += 1
        elif result.status == ImportStatus.WARNING:
            self.warning_count += 1
        elif result.status == ImportStatus.ERROR:
            self.error_count += 1
        else:
            self.skipped_count += 1
    
    def finish(self):
        self.finished_at = datetime.now()
    
    @property
    def duration_seconds(self) -> float:
        if self.finished_at:
            return (self.finished_at - self.started_at).total_seconds()
        return 0.0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "success_count": self.success_count,
            "warning_count": self.warning_count,
            "error_count": self.error_count,
            "skipped_count": self.skipped_count,
            "duration_seconds": self.duration_seconds,
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat() if self.finished_at else None,
        }


@dataclass
class ColumnMapping:
    """Mapeamento de colunas do arquivo para campos do sistema"""
    file_column: str
    system_field: str
    required: bool = False
    transformer: Optional[Callable[[Any], Any]] = None


class ServicoImportacao:
    """Serviço para importação de dados de arquivos CSV e Excel"""
    
    # Mapeamentos padrão para estudantes
    STUDENT_MAPPINGS = {
        'nome': ['nome', 'name', 'nome completo', 'name completo'],
        'email': ['email', 'e-mail', 'email contato', 'contato'],
        'curso': ['curso', 'course', 'matricula', 'matrícula'],
        'idade': ['idade', 'age', 'anos'],
        'telefone': ['telefone', 'phone', 'celular', 'fone'],
        'has_medical_report': ['laudo', 'has_medical_report', 'possui laudo', 'laudo médico'],
        'requires_attention': ['atencao', 'requires_attention', 'requer atencao', 'requer atenção'],
        'attention_reason': ['motivo', 'attention_reason', 'motivo atencao', 'razão'],
    }
    
    # Mapeamentos padrão para agendamentos
    APPOINTMENT_MAPPINGS = {
        'student_id': ['student_id', 'id_aluno', 'aluno_id', 'id estudante'],
        'data_hora': ['data_hora', 'datetime', 'data', 'data agendamento', 'horario'],
        'tipo': ['tipo', 'type', 'tipo atendimento'],
        'status': ['status', 'situacao', 'situação'],
        'observacoes': ['observacoes', 'notes', 'observações', 'obs'],
    }
    
    # Mapeamentos padrão para orientações
    ORIENTATION_MAPPINGS = {
        'student_id': ['student_id', 'id_aluno', 'aluno_id', 'id estudante'],
        'session_date': ['session_date', 'data', 'data sessao', 'data sessão'],
        'main_complaint': ['main_complaint', 'queixa', 'queixa principal', 'motivo'],
        'notes': ['notes', 'observacoes', 'observações', 'notas'],
        'themes': ['themes', 'temas', 'tema'],
    }
    
    def __init__(self):
        self._progress_callback: Optional[Callable[[int, int, str], None]] = None
    
    def set_progress_callback(self, callback: Callable[[int, int, str], None]):
        """Define callback para acompanhar progresso da importação"""
        self._progress_callback = callback
    
    def _report_progress(self, current: int, total: int, message: str):
        """Reporta progresso da importação"""
        if self._progress_callback:
            self._progress_callback(current, total, message)
    
    def _detect_column_mapping(self, headers: List[str], field_mappings: Dict[str, List[str]]) -> Dict[str, str]:
        """
        Detecta automaticamente o mapeamento de colunas baseado nos headers do arquivo.
        
        Returns:
            Dict mapeando nome do campo do sistema -> nome da coluna no arquivo
        """
        mapping = {}
        headers_lower = [h.lower().strip() for h in headers]
        
        for field, possible_names in field_mappings.items():
            for i, header in enumerate(headers_lower):
                if header in [n.lower() for n in possible_names]:
                    mapping[field] = headers[i]  # Usa o nome original da coluna
                    break
        
        return mapping
    
    def _transform_value(self, value: Any, field: str) -> Any:
        """Transforma um valor para o formato esperado pelo sistema"""
        if value is None or str(value).strip() == '':
            return None
        
        value_str = str(value).strip()
        
        # Transformações específicas por campo
        if field in ['has_medical_report', 'requires_attention']:
            # Converte para booleano
            return value_str.lower() in ['sim', 's', 'yes', 'y', '1', 'true', 'verdadeiro', 'x']
        
        if field in ['idade', 'student_id']:
            try:
                return int(float(value_str))
            except (ValueError, TypeError):
                return None
        
        if field in ['data_hora', 'session_date']:
            # Tenta vários formatos de data
            date_formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%d-%m-%Y',
            ]
            for fmt in date_formats:
                try:
                    return datetime.strptime(value_str, fmt)
                except ValueError:
                    continue
            return None
        
        return value_str
    
    def _validate_row(self, data: Dict[str, Any], required_fields: List[str]) -> List[str]:
        """Valida uma linha de dados e retorna lista de erros"""
        errors = []
        
        for field in required_fields:
            if field not in data or data[field] is None or str(data[field]).strip() == '':
                errors.append(f"Campo obrigatório '{field}' não preenchido")
        
        return errors
    
    def read_csv(self, file_path: str, encoding: str = 'utf-8') -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Lê um arquivo CSV e retorna headers e linhas.
        
        Returns:
            Tuple de (headers, rows)
        """
        rows = []
        headers = []
        
        # Tenta diferentes encodings
        encodings_to_try = [encoding, 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for enc in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
                    rows = list(reader)
                    logger.info(f"Arquivo CSV lido com encoding {enc}: {len(rows)} linhas")
                    break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.error(f"Erro ao ler CSV com encoding {enc}: {e}")
                continue
        
        return headers, rows
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Lê um arquivo Excel e retorna headers e linhas.
        
        Returns:
            Tuple de (headers, rows)
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl não está instalado. Instale com: pip install openpyxl")
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Seleciona a planilha
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Lê headers (primeira linha)
            headers = [str(cell.value or '') for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            # Lê dados
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, header in enumerate(headers):
                    if header:
                        row_dict[header] = row[i] if i < len(row) else None
                rows.append(row_dict)
            
            wb.close()
            logger.info(f"Arquivo Excel lido: {len(rows)} linhas")
            return headers, rows
            
        except Exception as e:
            logger.error(f"Erro ao ler Excel: {e}")
            raise
    
    def importar_estudantes(
        self,
        file_path: str,
        column_mapping: Optional[Dict[str, str]] = None,
        skip_duplicates: bool = True,
        update_existing: bool = False
    ) -> ImportReport:
        """
        Importa estudantes de um arquivo CSV ou Excel.
        
        Args:
            file_path: Caminho do arquivo
            column_mapping: Mapeamento customizado de colunas (opcional)
            skip_duplicates: Se deve pular registros duplicados
            update_existing: Se deve atualizar registros existentes
            
        Returns:
            ImportReport com resultados da importação
        """
        report = ImportReport()
        
        try:
            # Lê o arquivo
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                headers, rows = self.read_excel(file_path)
            else:
                raise ValueError(f"Formato de arquivo não suportado: {file_path}")
            
            report.total_rows = len(rows)
            
            # Detecta mapeamento de colunas
            if column_mapping:
                mapping = column_mapping
            else:
                mapping = self._detect_column_mapping(headers, self.STUDENT_MAPPINGS)
            
            logger.info(f"Mapeamento de colunas detectado: {mapping}")
            
            # Valida campos obrigatórios
            required_fields = ['nome']  # Nome é obrigatório
            missing_required = [f for f in required_fields if f not in mapping]
            if missing_required:
                raise ValueError(f"Colunas obrigatórias não encontradas: {missing_required}")
            
            # Processa cada linha
            for i, row in enumerate(rows, start=1):
                self._report_progress(i, len(rows), f"Processando linha {i} de {len(rows)}")
                
                # Extrai dados usando mapeamento
                student_data = {}
                for field, col_name in mapping.items():
                    if col_name in row:
                        student_data[field] = self._transform_value(row[col_name], field)
                
                # Valida linha
                errors = self._validate_row(student_data, required_fields)
                if errors:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message="; ".join(errors),
                        original_data=row
                    ))
                    continue
                
                # Verifica duplicatas
                if skip_duplicates or update_existing:
                    existing = self._find_existing_student(student_data.get('nome'), student_data.get('email'))
                    if existing:
                        if skip_duplicates and not update_existing:
                            report.add_result(ImportResult(
                                status=ImportStatus.SKIPPED,
                                row_number=i,
                                message=f"Estudante já existe: {student_data.get('nome')}",
                                data={"existing_id": existing['id']},
                                original_data=row
                            ))
                            continue
                        elif update_existing:
                            result = self._update_student(existing['id'], student_data)
                            report.add_result(ImportResult(
                                status=ImportStatus.SUCCESS if result['success'] else ImportStatus.ERROR,
                                row_number=i,
                                message=f"Estudante atualizado: {student_data.get('nome')}" if result['success'] else result.get('error', 'Erro ao atualizar'),
                                data={"id": existing['id']},
                                original_data=row
                            ))
                            continue
                
                # Cria novo estudante
                result = self._create_student(student_data)
                if result['success']:
                    report.add_result(ImportResult(
                        status=ImportStatus.SUCCESS,
                        row_number=i,
                        message=f"Estudante criado: {student_data.get('nome')}",
                        data={"id": result.get('id')},
                        original_data=row
                    ))
                else:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message=result.get('error', 'Erro ao criar estudante'),
                        original_data=row
                    ))
            
            report.finish()
            logger.info(f"Importação concluída: {report.to_dict()}")
            
        except Exception as e:
            logger.exception(f"Erro na importação: {e}")
            report.add_result(ImportResult(
                status=ImportStatus.ERROR,
                row_number=0,
                message=f"Erro fatal: {str(e)}"
            ))
            report.finish()
        
        return report
    
    def importar_agendamentos(
        self,
        file_path: str,
        column_mapping: Optional[Dict[str, str]] = None,
        skip_duplicates: bool = True
    ) -> ImportReport:
        """
        Importa agendamentos de um arquivo CSV ou Excel.
        
        Args:
            file_path: Caminho do arquivo
            column_mapping: Mapeamento customizado de colunas (opcional)
            skip_duplicates: Se deve pular registros duplicados
            
        Returns:
            ImportReport com resultados da importação
        """
        report = ImportReport()
        
        try:
            # Lê o arquivo
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                headers, rows = self.read_excel(file_path)
            else:
                raise ValueError(f"Formato de arquivo não suportado: {file_path}")
            
            report.total_rows = len(rows)
            
            # Detecta mapeamento de colunas
            if column_mapping:
                mapping = column_mapping
            else:
                mapping = self._detect_column_mapping(headers, self.APPOINTMENT_MAPPINGS)
            
            # Valida campos obrigatórios
            required_fields = ['student_id', 'data_hora']
            missing_required = [f for f in required_fields if f not in mapping]
            if missing_required:
                raise ValueError(f"Colunas obrigatórias não encontradas: {missing_required}")
            
            # Processa cada linha
            for i, row in enumerate(rows, start=1):
                self._report_progress(i, len(rows), f"Processando linha {i} de {len(rows)}")
                
                # Extrai dados usando mapeamento
                appointment_data = {}
                for field, col_name in mapping.items():
                    if col_name in row:
                        appointment_data[field] = self._transform_value(row[col_name], field)
                
                # Valida linha
                errors = self._validate_row(appointment_data, required_fields)
                if errors:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message="; ".join(errors),
                        original_data=row
                    ))
                    continue
                
                # Verifica se o estudante existe
                student_exists = self._check_student_exists(appointment_data['student_id'])
                if not student_exists:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message=f"Estudante não encontrado: ID {appointment_data['student_id']}",
                        original_data=row
                    ))
                    continue
                
                # Cria agendamento
                result = self._create_appointment(appointment_data)
                if result['success']:
                    report.add_result(ImportResult(
                        status=ImportStatus.SUCCESS,
                        row_number=i,
                        message=f"Agendamento criado para {appointment_data['data_hora']}",
                        data={"id": result.get('id')},
                        original_data=row
                    ))
                else:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message=result.get('error', 'Erro ao criar agendamento'),
                        original_data=row
                    ))
            
            report.finish()
            
        except Exception as e:
            logger.exception(f"Erro na importação: {e}")
            report.add_result(ImportResult(
                status=ImportStatus.ERROR,
                row_number=0,
                message=f"Erro fatal: {str(e)}"
            ))
            report.finish()
        
        return report
    
    def importar_orientacoes(
        self,
        file_path: str,
        column_mapping: Optional[Dict[str, str]] = None
    ) -> ImportReport:
        """
        Importa orientações de um arquivo JSON.
        
        Args:
            file_path: Caminho do arquivo JSON
            column_mapping: Mapeamento customizado de colunas (opcional)
            
        Returns:
            ImportReport com resultados da importação
        """
        import json
        
        report = ImportReport()
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Suporta tanto lista quanto objeto com array
            if isinstance(data, dict):
                if 'orientacoes' in data:
                    rows = data['orientacoes']
                elif 'data' in data:
                    rows = data['data']
                else:
                    rows = [data]
            else:
                rows = data
            
            report.total_rows = len(rows)
            
            # Processa cada orientação
            for i, orient_data in enumerate(rows, start=1):
                self._report_progress(i, len(rows), f"Processando orientação {i} de {len(rows)}")
                
                # Valida campos obrigatórios
                if not orient_data.get('student_id'):
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message="student_id é obrigatório",
                        original_data=orient_data
                    ))
                    continue
                
                # Cria orientação
                result = self._create_orientation(orient_data)
                if result['success']:
                    report.add_result(ImportResult(
                        status=ImportStatus.SUCCESS,
                        row_number=i,
                        message=f"Orientação criada",
                        data={"id": result.get('id')},
                        original_data=orient_data
                    ))
                else:
                    report.add_result(ImportResult(
                        status=ImportStatus.ERROR,
                        row_number=i,
                        message=result.get('error', 'Erro ao criar orientação'),
                        original_data=orient_data
                    ))
            
            report.finish()
            
        except Exception as e:
            logger.exception(f"Erro na importação: {e}")
            report.add_result(ImportResult(
                status=ImportStatus.ERROR,
                row_number=0,
                message=f"Erro fatal: {str(e)}"
            ))
            report.finish()
        
        return report
    
    def preview_import(self, file_path: str, entity_type: str = 'estudantes') -> Dict[str, Any]:
        """
        Gera preview dos dados que serão importados.
        
        Args:
            file_path: Caminho do arquivo
            entity_type: Tipo de entidade ('estudantes', 'agendamentos', 'orientacoes')
            
        Returns:
            Dict com headers, sample_data, detected_mapping, total_rows
        """
        try:
            # Lê o arquivo
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                headers, rows = self.read_excel(file_path)
            elif file_path.endswith('.json'):
                import json
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, list):
                    headers = list(data[0].keys()) if data else []
                    rows = data
                else:
                    headers = list(data.keys())
                    rows = [data]
            else:
                return {"success": False, "error": f"Formato não suportado: {file_path}"}
            
            # Detecta mapeamento
            mappings = {
                'estudantes': self.STUDENT_MAPPINGS,
                'agendamentos': self.APPOINTMENT_MAPPINGS,
                'orientacoes': self.ORIENTATION_MAPPINGS,
            }
            
            mapping = self._detect_column_mapping(headers, mappings.get(entity_type, {}))
            
            return {
                "success": True,
                "headers": headers,
                "sample_data": rows[:5],  # Primeiras 5 linhas
                "detected_mapping": mapping,
                "total_rows": len(rows),
                "file_type": os.path.splitext(file_path)[1]
            }
            
        except Exception as e:
            logger.exception(f"Erro ao gerar preview: {e}")
            return {"success": False, "error": str(e)}
    
    # Métodos auxiliares de banco de dados
    
    def _find_existing_student(self, nome: Optional[str], email: Optional[str]) -> Optional[Dict]:
        """Busca estudante existente por nome ou email"""
        if not nome and not email:
            return None
        
        try:
            connection = get_db_connection()
            cursor = connection.cursor(dictionary=True)
            
            query = "SELECT id_aluno as id, nome, email FROM aluno WHERE "
            params = []
            
            if nome and email:
                query += "(nome = %s OR email = %s)"
                params = [nome, email]
            elif nome:
                query += "nome = %s"
                params = [nome]
            else:
                query += "email = %s"
                params = [email]
            
            cursor.execute(query, params)
            result = cursor.fetchone()
            connection.close()
            
            return result
            
        except Exception as e:
            logger.error(f"Erro ao buscar estudante existente: {e}")
            return None
    
    def _check_student_exists(self, student_id: int) -> bool:
        """Verifica se um estudante existe"""
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            cursor.execute("SELECT 1 FROM aluno WHERE id_aluno = %s", (student_id,))
            result = cursor.fetchone()
            connection.close()
            return result is not None
        except Exception as e:
            logger.error(f"Erro ao verificar estudante: {e}")
            return False
    
    def _create_student(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Cria um novo estudante no banco"""
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            query = """
                INSERT INTO aluno (nome, email, curso, idade, phone, has_medical_report, requires_attention)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                data.get('nome'),
                data.get('email'),
                data.get('curso'),
                data.get('idade'),
                data.get('telefone'),
                data.get('has_medical_report', False),
                data.get('requires_attention', False)
            ))
            connection.commit()
            student_id = cursor.lastrowid
            connection.close()
            
            return {"success": True, "id": student_id}
            
        except Exception as e:
            logger.error(f"Erro ao criar estudante: {e}")
            return {"success": False, "error": str(e)}
    
    def _update_student(self, student_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """Atualiza um estudante existente"""
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            # Constrói query dinamicamente
            updates = []
            params = []
            
            field_mapping = {
                'nome': 'nome',
                'email': 'email',
                'curso': 'curso',
                'idade': 'idade',
                'telefone': 'phone',
                'has_medical_report': 'has_medical_report',
                'requires_attention': 'requires_attention',
            }
            
            for field, db_field in field_mapping.items():
                if field in data and data[field] is not None:
                    updates.append(f"{db_field} = %s")
                    params.append(data[field])
            
            if not updates:
                connection.close()
                return {"success": True, "message": "Nenhum campo para atualizar"}
            
            params.append(student_id)
            query = f"UPDATE aluno SET {', '.join(updates)} WHERE id_aluno = %s"
            
            cursor.execute(query, params)
            connection.commit()
            connection.close()
            
            return {"success": True}
            
        except Exception as e:
            logger.error(f"Erro ao atualizar estudante: {e}")
            return {"success": False, "error": str(e)}
    
    def _create_appointment(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Cria um novo agendamento no banco"""
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            query = """
                INSERT INTO agendamento (student_id, data_hora, tipo, status, observacoes, origem)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                data.get('student_id'),
                data.get('data_hora'),
                data.get('tipo', 'Atendimento'),
                data.get('status', 'agendado'),
                data.get('observacoes'),
                'importado'
            ))
            connection.commit()
            appointment_id = cursor.lastrowid
            connection.close()
            
            return {"success": True, "id": appointment_id}
            
        except Exception as e:
            logger.error(f"Erro ao criar agendamento: {e}")
            return {"success": False, "error": str(e)}
    
    def _create_orientation(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Cria uma nova orientação no banco"""
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            query = """
                INSERT INTO desktop_orientation (student_id, session_date, main_complaint, notes, themes)
                VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                data.get('student_id'),
                data.get('session_date'),
                data.get('main_complaint'),
                data.get('notes'),
                data.get('themes')
            ))
            connection.commit()
            orientation_id = cursor.lastrowid
            connection.close()
            
            return {"success": True, "id": orientation_id}
            
        except Exception as e:
            logger.error(f"Erro ao criar orientação: {e}")
            return {"success": False, "error": str(e)}


# Instância global para fácil acesso
servico_importacao = ServicoImportacao()
